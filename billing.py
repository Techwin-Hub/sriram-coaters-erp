import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import os # For checking file existence
import billing_db # Functions to interact with the billing tables in the database

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from num2words import num2words # For amount in words
from datetime import date # For default date

class BillingEntryPage(tk.Frame):
    EXCEL_FILE_NAME = "RANE 2025-2026.xlsx"
    EXCEL_SHEET_NAME = "BillingData"

    def __init__(self, master, main_app_title_var=None):
        super().__init__(master)
        self.main_app_title_var = main_app_title_var
        self.configure(bg="white")
        self.line_items_data = []
        self.next_line_no = 1
        self.selected_invoice_no = None
        self.edit_mode = False
        self.editing_invoice_no = None

        # Register Courier font if not already registered (globally for reportlab)
        # Also, set initial invoice number
        try:
            pdfmetrics.getFont('Courier')
        except KeyError: # Font not yet registered
            try:
                pdfmetrics.registerFont(TTFont('Courier', 'Courier.ttf'))
                pdfmetrics.registerFont(TTFont('Courier-Bold', 'Courier-Bold.ttf')) # If you have bold variant
            except Exception as e:
                print(f"Warning: Could not register Courier font for PDF: {e}. Using Helvetica.")


        header_frame = ttk.LabelFrame(self, text="Invoice Details", padding=(10, 5))
        header_frame.pack(padx=20, pady=10, fill="x")

        # Row 0
        ttk.Label(header_frame, text="Invoice Number:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.invoice_no_var = tk.StringVar()
        self.invoice_no_var.set(self._generate_next_invoice_no()) # Set initial invoice number
        self.invoice_no_entry = ttk.Entry(header_frame, textvariable=self.invoice_no_var, width=30)
        self.invoice_no_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(header_frame, text="Date (YYYY-MM-DD):").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.date_var = tk.StringVar(value=date.today().isoformat())
        self.date_entry = ttk.Entry(header_frame, textvariable=self.date_var, width=30)
        self.date_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        # Row 1
        ttk.Label(header_frame, text="Customer Name:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.customer_name_var = tk.StringVar()
        self.customer_name_entry = ttk.Entry(header_frame, textvariable=self.customer_name_var, width=30)
        self.customer_name_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(header_frame, text="Payment Method:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.payment_method_var = tk.StringVar()
        self.payment_method_combo = ttk.Combobox(header_frame, textvariable=self.payment_method_var,
                                                 values=["Cash", "Card", "Online Transfer", "Cheque"], width=27)
        self.payment_method_combo.grid(row=1, column=3, padx=5, pady=5, sticky="ew")
        self.payment_method_combo.set("Cash")

        # Row 2
        ttk.Label(header_frame, text="GST % (e.g., 18):").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.gst_percentage_var = tk.StringVar()
        self.gst_percentage_entry = ttk.Entry(header_frame, textvariable=self.gst_percentage_var, width=30)
        self.gst_percentage_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(header_frame, text="PO Number:").grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.po_number_var = tk.StringVar()
        self.po_number_entry = ttk.Entry(header_frame, textvariable=self.po_number_var, width=30)
        self.po_number_entry.grid(row=2, column=3, padx=5, pady=5, sticky="ew")

        # Row 3 - GRN Dates
        ttk.Label(header_frame, text="GRN Date From:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.grn_date_from_var = tk.StringVar()
        self.grn_date_from_entry = ttk.Entry(header_frame, textvariable=self.grn_date_from_var, width=30)
        self.grn_date_from_entry.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(header_frame, text="GRN Date To:").grid(row=3, column=2, padx=5, pady=5, sticky="w")
        self.grn_date_to_var = tk.StringVar()
        self.grn_date_to_entry = ttk.Entry(header_frame, textvariable=self.grn_date_to_var, width=30)
        self.grn_date_to_entry.grid(row=3, column=3, padx=5, pady=5, sticky="ew")

        header_frame.columnconfigure(1, weight=1)
        header_frame.columnconfigure(3, weight=1)

        # ... (rest of __init__ method: line item entry, current items tree, history, summary/actions)
        line_item_entry_frame = ttk.LabelFrame(self, text="Add Line Item", padding=(10, 5))
        line_item_entry_frame.pack(padx=20, pady=10, fill="x")

        ttk.Label(line_item_entry_frame, text="Item Description:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.item_desc_var = tk.StringVar()
        self.item_desc_entry = ttk.Entry(line_item_entry_frame, textvariable=self.item_desc_var, width=40)
        self.item_desc_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(line_item_entry_frame, text="Part No:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.part_no_var = tk.StringVar()
        self.part_no_entry = ttk.Entry(line_item_entry_frame, textvariable=self.part_no_var, width=20)
        self.part_no_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        ttk.Label(line_item_entry_frame, text="HSN Code:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.hsn_code_var = tk.StringVar()
        self.hsn_code_entry = ttk.Entry(line_item_entry_frame, textvariable=self.hsn_code_var, width=40)
        self.hsn_code_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(line_item_entry_frame, text="Quantity:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.quantity_var = tk.StringVar()
        self.quantity_entry = ttk.Entry(line_item_entry_frame, textvariable=self.quantity_var, width=10)
        self.quantity_entry.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.quantity_var.trace_add("write", self.calculate_line_item_amount)

        ttk.Label(line_item_entry_frame, text="Rate:").grid(row=1, column=4, padx=5, pady=5, sticky="w")
        self.rate_var = tk.StringVar()
        self.rate_entry = ttk.Entry(line_item_entry_frame, textvariable=self.rate_var, width=10)
        self.rate_entry.grid(row=1, column=5, padx=5, pady=5, sticky="w")
        self.rate_var.trace_add("write", self.calculate_line_item_amount)

        ttk.Label(line_item_entry_frame, text="Amount:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.line_item_amount_var = tk.StringVar(value="0.00")
        self.line_item_amount_label = ttk.Label(line_item_entry_frame, textvariable=self.line_item_amount_var, width=15, anchor="w")
        self.line_item_amount_label.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        add_item_button = ttk.Button(line_item_entry_frame, text="Add Item to Invoice",
                                     command=self.add_line_item_to_treeview, style="Accent.TButton")
        add_item_button.grid(row=2, column=2, columnspan=4, padx=5, pady=10, sticky="e")

        line_item_entry_frame.columnconfigure(1, weight=1)
        line_item_entry_frame.columnconfigure(3, weight=1)

        line_items_display_frame = ttk.LabelFrame(self, text="Current Invoice Items", padding=(10, 5))
        line_items_display_frame.pack(padx=20, pady=10, fill="both", expand=True)

        columns = ("line_no", "item_desc", "part_no", "hsn_code", "qty", "rate", "amount")
        self.tree = ttk.Treeview(line_items_display_frame, columns=columns, show="headings", height=5)
        self.tree.heading("line_no", text="Line #")
        self.tree.heading("item_desc", text="Item Desc")
        self.tree.heading("part_no", text="Part No")
        self.tree.heading("hsn_code", text="HSN")
        self.tree.heading("qty", text="Qty")
        self.tree.heading("rate", text="Rate")
        self.tree.heading("amount", text="Amount")
        self.tree.column("line_no", width=50, anchor="center", stretch=tk.NO)
        self.tree.column("item_desc", width=250, anchor="w")
        self.tree.column("part_no", width=100, anchor="w", stretch=tk.NO)
        self.tree.column("hsn_code", width=80, anchor="w", stretch=tk.NO)
        self.tree.column("qty", width=70, anchor="e", stretch=tk.NO)
        self.tree.column("rate", width=80, anchor="e", stretch=tk.NO)
        self.tree.column("amount", width=100, anchor="e", stretch=tk.NO)
        tree_scrollbar = ttk.Scrollbar(line_items_display_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scrollbar.pack(side="right", fill="y")

        history_frame = ttk.LabelFrame(self, text="Billing History", padding=(10, 5))
        history_frame.pack(padx=20, pady=10, fill="both", expand=True)

        search_controls_frame = ttk.Frame(history_frame)
        search_controls_frame.pack(fill="x", pady=(0,5))
        ttk.Label(search_controls_frame, text="Search by Invoice No:").pack(side="left", padx=(0,5))
        self.search_invoice_no_var = tk.StringVar()
        self.search_invoice_no_entry = ttk.Entry(search_controls_frame, textvariable=self.search_invoice_no_var, width=25)
        self.search_invoice_no_entry.pack(side="left", padx=5)
        search_button = ttk.Button(search_controls_frame, text="Search", command=self._search_invoices)
        search_button.pack(side="left", padx=5)
        clear_search_button = ttk.Button(search_controls_frame, text="Show All", command=self._load_billing_history)
        clear_search_button.pack(side="left", padx=5)

        history_columns = ("invoice_no", "date", "customer_name", "total_amount", "gst_percentage", "payment_method") # PO Number could be added here
        self.history_tree = ttk.Treeview(history_frame, columns=history_columns, show="headings", height=6)
        self.history_tree.heading("invoice_no", text="Invoice No")
        self.history_tree.heading("date", text="Date")
        self.history_tree.heading("customer_name", text="Customer Name")
        self.history_tree.heading("total_amount", text="Total Amount")
        self.history_tree.heading("gst_percentage", text="GST %")
        self.history_tree.heading("payment_method", text="Payment Method")
        self.history_tree.column("invoice_no", width=100, anchor="w")
        self.history_tree.column("date", width=80, anchor="center")
        self.history_tree.column("customer_name", width=200, anchor="w")
        self.history_tree.column("total_amount", width=100, anchor="e")
        self.history_tree.column("gst_percentage", width=60, anchor="e")
        self.history_tree.column("payment_method", width=100, anchor="w")
        history_scrollbar = ttk.Scrollbar(history_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=history_scrollbar.set)
        self.history_tree.pack(side="left", fill="both", expand=True)
        history_scrollbar.pack(side="right", fill="y")
        self.history_tree.bind("<<TreeviewSelect>>", self._on_history_select)

        history_actions_frame = ttk.Frame(history_frame)
        history_actions_frame.pack(fill="x", pady=(5,0))
        self.edit_invoice_button = ttk.Button(history_actions_frame, text="Edit Selected", command=self._edit_selected_invoice, state="disabled")
        self.edit_invoice_button.pack(side="left", padx=5)
        self.delete_invoice_button = ttk.Button(history_actions_frame, text="Delete Selected", command=self._delete_selected_invoice, state="disabled")
        self.delete_invoice_button.pack(side="left", padx=5)
        self.export_pdf_button = ttk.Button(history_actions_frame, text="Export to PDF", command=self._export_selected_to_pdf, state="disabled")
        self.export_pdf_button.pack(side="left", padx=5)

        summary_actions_frame = ttk.Frame(self, padding=(10,5))
        summary_actions_frame.pack(padx=20, pady=(0,10), fill="x")
        ttk.Label(summary_actions_frame, text="Total Invoice Amount:").pack(side="left", padx=(0,5))
        self.total_invoice_amount_var = tk.StringVar(value="0.00")
        total_invoice_amount_label = ttk.Label(summary_actions_frame, textvariable=self.total_invoice_amount_var, font=("Arial", 12, "bold"))
        total_invoice_amount_label.pack(side="left")
        self.save_invoice_button = ttk.Button(summary_actions_frame, text="Save Invoice", command=self.save_invoice, style="Accent.TButton")
        self.save_invoice_button.pack(side="right", padx=5)

        self._load_billing_history()

    def _generate_next_invoice_no(self) -> str:
        last_invoice_no_str = billing_db.get_last_invoice_no()
        if last_invoice_no_str:
            try:
                # Attempt to extract numeric part for incrementing
                match = None
                for i, char_val in enumerate(last_invoice_no_str):
                    if char_val.isdigit():
                        match = i
                        break

                if match is not None:
                    prefix = last_invoice_no_str[:match]
                    numeric_part_str = last_invoice_no_str[match:]
                    if numeric_part_str.isdigit():
                        next_num = int(numeric_part_str) + 1
                        return f"{prefix}{str(next_num).zfill(len(numeric_part_str))}"

                if last_invoice_no_str.isdigit():
                    next_num = int(last_invoice_no_str) + 1
                    return str(next_num).zfill(len(last_invoice_no_str))

                print(f"Warning: Could not reliably increment invoice number '{last_invoice_no_str}'. Defaulting.")
                return "001"

            except ValueError:
                print(f"Warning: Could not parse last invoice number '{last_invoice_no_str}' for incrementing. Defaulting.")
                return "001"
        else:
            return "001"

    def _on_history_select(self, event):
        selected_items = self.history_tree.selection()
        if selected_items:
            selected_item = selected_items[0]
            self.selected_invoice_no = self.history_tree.item(selected_item, "values")[0]
            self.edit_invoice_button.config(state="normal")
            self.delete_invoice_button.config(state="normal")
            self.export_pdf_button.config(state="normal")
        else:
            self.selected_invoice_no = None
            self.edit_invoice_button.config(state="disabled")
            self.delete_invoice_button.config(state="disabled")
            self.export_pdf_button.config(state="disabled")

    def _edit_selected_invoice(self):
        if not self.selected_invoice_no:
            messagebox.showwarning("No Selection", "Please select an invoice from the history to edit.")
            return

        self.edit_mode = True
        self.editing_invoice_no = self.selected_invoice_no

        header_data = billing_db.get_invoice_header_by_no(self.editing_invoice_no)
        line_items = billing_db.get_line_items_for_invoice(self.editing_invoice_no)

        if not header_data:
            messagebox.showerror("Error", f"Could not retrieve invoice details for {self.editing_invoice_no}.")
            self._clear_form_and_reset()
            return

        self.invoice_no_var.set(header_data.get('invoice_no', ''))
        self.date_var.set(header_data.get('date', ''))
        self.customer_name_var.set(header_data.get('customer_name', ''))
        self.payment_method_var.set(header_data.get('payment_method', 'Cash'))
        self.gst_percentage_var.set(str(header_data.get('gst_percentage', '')))
        self.po_number_var.set(header_data.get('po_number', '')) # Populate PO Number
        self.grn_date_from_var.set(header_data.get('grn_date_from', '')) # Populate GRN Date From
        self.grn_date_to_var.set(header_data.get('grn_date_to', ''))     # Populate GRN Date To
        self.invoice_no_entry.config(state="readonly")

        self.line_items_data.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.next_line_no = 1
        if line_items:
            for item_db in line_items:
                item_ui = {
                    "line_no": item_db['line_no'], "item_description": item_db['item_description'],
                    "part_no": item_db['part_no'], "hsn_code": item_db['hsn_code'],
                    "quantity": float(item_db['quantity']), "rate": float(item_db['rate']),
                    "amount": float(item_db['amount'])
                }
                self.line_items_data.append(item_ui)
                self.tree.insert("", tk.END, values=(
                    item_ui["line_no"], item_ui["item_description"], item_ui["part_no"],
                    item_ui["hsn_code"], f"{item_ui['quantity']:.2f}",
                    f"{item_ui['rate']:.2f}", f"{item_ui['amount']:.2f}"
                ))
                self.next_line_no = max(self.next_line_no, item_ui["line_no"] + 1)

        self.update_total_invoice_amount()
        self.save_invoice_button.config(text="Update Invoice")
        if self.main_app_title_var:
            self.main_app_title_var.set(f"Edit Invoice: {self.editing_invoice_no}")
        # Ensure invoice number entry is focused for new/edit
        self.invoice_no_entry.focus_set()
        if self.edit_mode: # If editing, move focus to date perhaps, as invoice_no is readonly
            self.date_entry.focus_set()


    def _delete_selected_invoice(self):
        if not self.selected_invoice_no:
            messagebox.showwarning("No Selection", "Please select an invoice from the history to delete.")
            return

        confirm = messagebox.askyesno("Confirm Delete",
                                     f"Are you sure you want to delete invoice {self.selected_invoice_no}?\n"
                                     "This will also delete all associated line items.")
        if confirm:
            if billing_db.delete_invoice_header(self.selected_invoice_no):
                messagebox.showinfo("Success", f"Invoice {self.selected_invoice_no} deleted successfully.")
                self._load_billing_history()
            else:
                messagebox.showerror("Error", f"Failed to delete invoice {self.selected_invoice_no}.")

        self.selected_invoice_no = None
        self.edit_invoice_button.config(state="disabled")
        self.delete_invoice_button.config(state="disabled")
        self.export_pdf_button.config(state="disabled")

    def _export_selected_to_pdf(self):
        if not self.selected_invoice_no:
            messagebox.showwarning("No Selection", "Please select an invoice from the history to export.")
            return

        header_data = billing_db.get_invoice_header_by_no(self.selected_invoice_no)
        line_items_data = billing_db.get_line_items_for_invoice(self.selected_invoice_no)

        if not header_data:
            messagebox.showerror("Error", f"Could not retrieve data for invoice {self.selected_invoice_no}.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF documents", "*.pdf")],
            initialfile=f"Invoice_{self.selected_invoice_no}.pdf",
            title="Save Invoice PDF"
        )
        if not file_path:
            return

        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter,
                                    rightMargin=0.5*inch, leftMargin=0.5*inch,
                                    topMargin=0.5*inch, bottomMargin=0.5*inch)

            styles = getSampleStyleSheet()
            # Attempt to use Courier, fallback to Helvetica if Courier not found/registered
            try:
                pdfmetrics.getFont('Courier') # Check if Courier is available
                courier_font_name = 'Courier'
            except KeyError:
                courier_font_name = 'Helvetica' # Fallback font
                print("Warning: Courier font not found, using Helvetica for PDF.")

            courier_style = ParagraphStyle('Courier', parent=styles['Normal'], fontName=courier_font_name, fontSize=10, leading=12)
            courier_bold_style = ParagraphStyle('CourierBold', parent=courier_style, fontName=courier_font_name + '-Bold' if courier_font_name == 'Courier' else 'Helvetica-Bold')
            company_name_style = ParagraphStyle('CompName', parent=courier_style, alignment=1, fontSize=12, fontName=courier_font_name + '-Bold' if courier_font_name == 'Courier' else 'Helvetica-Bold')
            centered_courier_style = ParagraphStyle('CenteredCourier', parent=courier_style, alignment=1)


            story = []

            # Header
            header_text_1 = f"SSI NO. 332213072/26-12-03{'LABOUR BILL INVOICE'.center(30)}{'MOB: 9443410161'.rjust(25)}"
            header_text_2 = f"GSTin NO: 33ALQPK2156A1zg{'E-MAIL: kasinatharm@yahoo.in'.rjust(55)}" # Adjusted spacing
            story.append(Paragraph(header_text_1, courier_style))
            story.append(Paragraph(header_text_2, courier_style))
            story.append(Spacer(1, 0.05*inch))
            story.append(Paragraph("SRIRAM     COATERS", company_name_style))
            story.append(Paragraph("AMBAL NAGAR, BOOTHAKUDI VILLAGE, VIRALIMALAI-621316", centered_courier_style))
            story.append(Paragraph("*"*78, centered_courier_style)) # Approx 78 chars for 6.5 inch width at 10cpi

            # Invoice Info (Inv No, Date)
            inv_no_str = f"INVOICE NO: {header_data.get('invoice_no', 'N/A')}"
            date_str = f"* INV.DATE * {header_data.get('date', 'N/A')}"
            # Attempt to align date to the right. Max chars approx 78.
            padding_needed = 78 - len(inv_no_str) - len(date_str)
            inv_info_line = f"{inv_no_str}{' ' * max(0, padding_needed)}{date_str}"
            story.append(Paragraph(inv_info_line, courier_style))
            story.append(Paragraph("*"*78, centered_courier_style))

            # Customer/PO Info
            # Using a table for better alignment of left and right content
            customer_gstin_placeholder = "33AAACR3147C1ZY" # Placeholder for actual customer GSTIN
            left_col_text = f"""TO<br/>
                                THE MANAGER,<br/>
                                M/s {header_data.get('customer_name', 'N/A')},<br/>
                                BOOTHAKUDI VILLAGE, VIRALIMALAI-621316.<br/>
                                <br/>
                                GSTin NO: {customer_gstin_placeholder}""" # Replace with actual data if available

            right_col_text = f"""*<br/>
                                 *PONO:     {header_data.get('po_number', 'N/A')}<br/>
                                 *<br/>
                                 *<br/>
                                 *"""

            customer_po_data = [[Paragraph(left_col_text, courier_style), Paragraph(right_col_text, courier_style)]]
            table_cust_po = Table(customer_po_data, colWidths=[4.5*inch, 2.8*inch]) # Adjusted total width
            table_cust_po.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(table_cust_po)
            story.append(Paragraph("*"*78, centered_courier_style))


            # Line Items Table
            table_data = [[
                Paragraph("SNO", courier_bold_style),
                Paragraph("ITEM DESCRIPTION", courier_bold_style),
                Paragraph("PART NO", courier_bold_style),
                Paragraph("HSN CODE", courier_bold_style),
                Paragraph("QATY", courier_bold_style),
                Paragraph("RATE", courier_bold_style),
                Paragraph("AMOUNT", courier_bold_style)
            ]]

            for item in line_items_data:
                table_data.append([
                    Paragraph(str(item['line_no']), courier_style),
                    Paragraph(item['item_description'], courier_style),
                    Paragraph(item['part_no'] if item['part_no'] else '', courier_style),
                    Paragraph(item['hsn_code'] if item['hsn_code'] else '', courier_style),
                    Paragraph(f"{item['quantity']:.2f}", courier_style),
                    Paragraph(f"{item['rate']:.2f}", courier_style),
                    Paragraph(f"{item['amount']:.2f}", courier_style)
                ])

            # Approximate column widths for Courier 10pt on 6.5 inch usable width (78 chars)
            # SNO(4) * DESC(30) * PART(10) * HSN(8) * QTY(8) * RATE(8) * AMOUNT(10) = 78
            col_widths_items = [0.4*inch, 3.0*inch, 1.0*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.0*inch]
            line_items_table = Table(table_data, colWidths=col_widths_items, repeatRows=1)

            line_items_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,0), 'CENTER'), # Header center
                ('FONTNAME', (0,0), (-1,0), courier_font_name + '-Bold' if courier_font_name == 'Courier' else 'Helvetica-Bold'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                # Data cells alignment
                ('ALIGN', (0,1), (0,-1), 'CENTER'), # SNO
                ('ALIGN', (1,1), (1,-1), 'LEFT'),   # Description
                ('ALIGN', (2,1), (3,-1), 'LEFT'),  # Part No, HSN
                ('ALIGN', (4,1), (-1,-1), 'RIGHT'), # Qty, Rate, Amount
            ]))
            story.append(line_items_table)
            story.append(Paragraph("*"*78, centered_courier_style))


            # Totals section
            total_val = float(header_data.get('total_amount', 0.0))
            gst_perc = float(header_data.get('gst_percentage', 0.0))

            # Assuming GST is split equally for CGST and SGST
            cgst_perc = gst_perc / 2
            sgst_perc = gst_perc / 2
            cgst_val = (total_val * cgst_perc) / 100
            sgst_val = (total_val * sgst_perc) / 100
            grand_total_val = total_val + cgst_val + sgst_val

            totals_data = [
                ['', Paragraph('TOTAL', courier_style), Paragraph(f"{total_val:.2f}", courier_style)],
                ['', Paragraph(f"CGST {cgst_perc:.1f}%", courier_style), Paragraph(f"{cgst_val:.2f}", courier_style)],
                ['', Paragraph(f"SGST {sgst_perc:.1f}%", courier_style), Paragraph(f"{sgst_val:.2f}", courier_style)],
                ['', Paragraph('GRAND TOTAL', courier_bold_style), Paragraph(f"{grand_total_val:.2f}", courier_bold_style)],
            ]
            totals_table = Table(totals_data, colWidths=[4.8*inch, 1.5*inch, 1.5*inch]) # Adjusted widths
            totals_table.setStyle(TableStyle([
                ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                 ('TOPPADDING', (0,0), (-1,-1), 1),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ]))
            story.append(totals_table)
            story.append(Paragraph("*"*78, centered_courier_style))

            # Amount in Words
            amount_words_prefix = "AMOUNT IN WORDS:"
            # Use num2words, ensure it handles decimals correctly for Indian currency or format as needed
            # For example, get integer part and decimal part separately if num2words has issues with paisa
            rupees = int(grand_total_val)
            paisa = int(round((grand_total_val - rupees) * 100))

            amount_words_str = f"{num2words(rupees, lang='en_IN').upper()} RUPEES"
            if paisa > 0:
                amount_words_str += f" AND {num2words(paisa, lang='en_IN').upper()} PAISA"
            amount_words_str += " ONLY"

            story.append(Paragraph(amount_words_prefix, courier_style))
            story.append(Paragraph(amount_words_str, courier_style)) # This might need wrapping if too long
            story.append(Paragraph("*"*78, centered_courier_style))


            # GRN Listing Date
            grn_from = header_data.get('grn_date_from', 'N/A')
            grn_to = header_data.get('grn_date_to', 'N/A')
            story.append(Paragraph(f"GRN LISTING DATE BETWEEN<br/>{grn_from} TO {grn_to}", courier_style))
            story.append(Paragraph("*"*78, centered_courier_style))

            # Footer
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph("FOR SRIRAM COATERS", ParagraphStyle('Footer', parent=courier_style, alignment=2))) # Right align

            doc.build(story)
            messagebox.showinfo("Success", f"PDF saved to {file_path}")
        except Exception as e:
            messagebox.showerror("PDF Export Error", f"Failed to generate PDF: {e}")

    def add_line_item_to_treeview(self):
        item_desc = self.item_desc_var.get()
        quantity_str = self.quantity_var.get()
        rate_str = self.rate_var.get()

        if not item_desc or not quantity_str or not rate_str:
            messagebox.showwarning("Missing Information", "Item Description, Quantity, and Rate are required to add an item.")
            return

        try:
            qty = float(quantity_str)
            rate = float(rate_str)
            if qty <= 0 or rate < 0:
                 messagebox.showwarning("Invalid Value", "Quantity must be positive and Rate must be non-negative.")
                 return
        except ValueError:
            messagebox.showerror("Invalid Input", "Quantity and Rate must be valid numbers.")
            return

        current_line_no = self.next_line_no

        item_data = {
            "line_no": current_line_no, "item_description": item_desc,
            "part_no": self.part_no_var.get(), "hsn_code": self.hsn_code_var.get(),
            "quantity": qty, "rate": rate, "amount": qty * rate
        }
        self.line_items_data.append(item_data)
        self.tree.insert("", tk.END, values=(
            item_data["line_no"], item_data["item_description"], item_data["part_no"],
            item_data["hsn_code"], f"{item_data['quantity']:.2f}",
            f"{item_data['rate']:.2f}", f"{item_data['amount']:.2f}"
        ))
        self.next_line_no += 1
        self.update_total_invoice_amount()

        self.item_desc_var.set("")
        self.part_no_var.set("")
        self.hsn_code_var.set("")
        self.quantity_var.set("")
        self.rate_var.set("")
        self.line_item_amount_var.set("0.00")
        self.item_desc_entry.focus()

    def calculate_line_item_amount(self, *args):
        try:
            qty = float(self.quantity_var.get() or 0)
            rate = float(self.rate_var.get() or 0)
            amount = qty * rate
            self.line_item_amount_var.set(f"{amount:.2f}")
        except ValueError:
            pass

    def update_total_invoice_amount(self):
        total = sum(item['amount'] for item in self.line_items_data)
        self.total_invoice_amount_var.set(f"{total:.2f}")

    def _save_line_items_to_excel(self, invoice_no, line_items):
        try:
            if os.path.exists(self.EXCEL_FILE_NAME):
                workbook = openpyxl.load_workbook(self.EXCEL_FILE_NAME)
                if self.EXCEL_SHEET_NAME in workbook.sheetnames:
                    sheet = workbook[self.EXCEL_SHEET_NAME]
                else:
                    sheet = workbook.create_sheet(title=self.EXCEL_SHEET_NAME)
                    headers = ["Invoice No", "Line No", "Item Description", "Part No", "HSN Code", "Quantity", "Rate", "Amount"]
                    sheet.append(headers)
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = self.EXCEL_SHEET_NAME
                headers = ["Invoice No", "Line No", "Item Description", "Part No", "HSN Code", "Quantity", "Rate", "Amount"]
                sheet.append(headers)

            for item in line_items:
                row_data = [
                    invoice_no, item['line_no'], item['item_description'], item['part_no'],
                    item['hsn_code'], item['quantity'], item['rate'], item['amount']
                ]
                sheet.append(row_data)

            workbook.save(self.EXCEL_FILE_NAME)
            return True
        except InvalidFileException:
            messagebox.showerror("Excel Error", f"File '{self.EXCEL_FILE_NAME}' is corrupted or not a valid Excel file. Please check or delete the file and try again.")
            return False
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to save line items to Excel: {e}")
            return False

    def save_invoice(self):
        if self.edit_mode:
            self._update_invoice_logic()
        else:
            self._create_new_invoice_logic()

    def _create_new_invoice_logic(self):
        invoice_no = self.invoice_no_var.get().strip()
        invoice_date = self.date_var.get().strip()
        customer_name = self.customer_name_var.get().strip()
        payment_method = self.payment_method_var.get()
        gst_percentage_str = self.gst_percentage_var.get().strip()
        po_number = self.po_number_var.get().strip()
        grn_date_from = self.grn_date_from_var.get().strip()
        grn_date_to = self.grn_date_to_var.get().strip()


        if not invoice_no or not invoice_date or not customer_name:
            messagebox.showerror("Validation Error", "Invoice Number, Date, and Customer Name are required.")
            return
        if not self.line_items_data:
            messagebox.showerror("Validation Error", "Cannot save an empty invoice. Please add line items.")
            return
        try:
            gst_percentage = float(gst_percentage_str) if gst_percentage_str else 0.0
        except ValueError:
            messagebox.showerror("Validation Error", "GST % must be a valid number.")
            return

        total_invoice_amount = sum(item['amount'] for item in self.line_items_data)
        header_data = {
            'invoice_no': invoice_no, 'date': invoice_date, 'customer_name': customer_name,
            'total_amount': total_invoice_amount, 'gst_percentage': gst_percentage,
            'payment_method': payment_method, 'po_number': po_number,
            'grn_date_from': grn_date_from if grn_date_from else None,
            'grn_date_to': grn_date_to if grn_date_to else None
        }

        save_header_result = billing_db.add_invoice_header(header_data)

        if save_header_result is not True:
            if save_header_result == "DUPLICATE":
                messagebox.showerror("Duplicate Invoice", f"Invoice number '{invoice_no}' already exists. Please use a different invoice number.")
            else:
                messagebox.showerror("Database Error", f"Failed to save invoice header for {invoice_no}. Please check application logs for details or ensure database is correctly set_up.")
            return


        line_items_for_db = [{'invoice_no': invoice_no, **item} for item in self.line_items_data]
        if not billing_db.add_invoice_line_items_batch(line_items_for_db):
            messagebox.showerror("Database Error", f"Failed to save line items for {invoice_no}.")
            billing_db.delete_invoice_header(invoice_no)
            return

        excel_save_success = self._save_line_items_to_excel(invoice_no, self.line_items_data)
        if excel_save_success:
            messagebox.showinfo("Success", f"Invoice {invoice_no} saved successfully to database and Excel!")
        else:
            messagebox.showwarning("Partial Success", f"Invoice {invoice_no} saved to database, but FAILED to save to Excel.")

        self._clear_form_and_reset()
        self._load_billing_history()

    def _update_invoice_logic(self):
        if not self.editing_invoice_no:
            messagebox.showerror("Error", "No invoice selected for update.")
            return

        invoice_date = self.date_var.get().strip()
        customer_name = self.customer_name_var.get().strip()
        payment_method = self.payment_method_var.get()
        gst_percentage_str = self.gst_percentage_var.get().strip()
        po_number = self.po_number_var.get().strip()
        grn_date_from = self.grn_date_from_var.get().strip()
        grn_date_to = self.grn_date_to_var.get().strip()

        if not invoice_date or not customer_name:
            messagebox.showerror("Validation Error", "Date and Customer Name are required.")
            return
        if not self.line_items_data:
            messagebox.showerror("Validation Error", "Invoice must have at least one line item.")
            return
        try:
            gst_percentage = float(gst_percentage_str) if gst_percentage_str else 0.0
        except ValueError:
            messagebox.showerror("Validation Error", "GST % must be a valid number.")
            return

        total_invoice_amount = sum(item['amount'] for item in self.line_items_data)
        updated_header_data = {
            'date': invoice_date, 'customer_name': customer_name,
            'total_amount': total_invoice_amount, 'gst_percentage': gst_percentage,
            'payment_method': payment_method, 'po_number': po_number,
            'grn_date_from': grn_date_from if grn_date_from else None,
            'grn_date_to': grn_date_to if grn_date_to else None
        }

        if not billing_db.update_invoice_header(self.editing_invoice_no, updated_header_data):
            messagebox.showerror("Database Error", f"Failed to update invoice header for {self.editing_invoice_no}.")
            return

        if not billing_db.delete_line_items_for_invoice(self.editing_invoice_no):
            messagebox.showerror("Database Error", f"Failed to delete old line items for {self.editing_invoice_no}. Update aborted.")
            return

        line_items_for_db = [{'invoice_no': self.editing_invoice_no, **item} for item in self.line_items_data]
        if not billing_db.add_invoice_line_items_batch(line_items_for_db):
            messagebox.showerror("Database Error", f"Failed to save updated line items for {self.editing_invoice_no}.")
            return

        messagebox.showinfo("Invoice Updated",
                            f"Invoice {self.editing_invoice_no} updated in database. "
                            "Excel file was NOT updated with these changes.")

        self._clear_form_and_reset()
        self._load_billing_history()

    def _clear_form_and_reset(self):
        self.invoice_no_var.set(self._generate_next_invoice_no()) # Set next suggested invoice number
        self.date_var.set(date.today().isoformat())
        self.customer_name_var.set("")
        self.payment_method_combo.set("Cash")
        self.gst_percentage_var.set("")
        self.po_number_var.set("") # Clear PO Number
        self.grn_date_from_var.set("") # Clear GRN Date From
        self.grn_date_to_var.set("")   # Clear GRN Date To

        self.line_items_data.clear()
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.next_line_no = 1

        self.item_desc_var.set("")
        self.part_no_var.set("")
        self.hsn_code_var.set("")
        self.quantity_var.set("")
        self.rate_var.set("")
        self.line_item_amount_var.set("0.00")
        self.total_invoice_amount_var.set("0.00")

        self.edit_mode = False
        self.editing_invoice_no = None
        self.invoice_no_entry.config(state="normal")
        self.save_invoice_button.config(text="Save Invoice")

        if self.main_app_title_var:
            self.main_app_title_var.set("Billing Management")

        self.invoice_no_entry.focus()

    def _load_billing_history(self, headers_to_load=None):
        for i in self.history_tree.get_children():
            self.history_tree.delete(i)

        self.selected_invoice_no = None
        self.edit_invoice_button.config(state="disabled")
        self.delete_invoice_button.config(state="disabled")
        self.export_pdf_button.config(state="disabled")


        if headers_to_load is None:
            invoice_headers = billing_db.get_all_invoice_headers()
        else:
            invoice_headers = headers_to_load

        if invoice_headers:
            for header in invoice_headers:
                # PO Number and GRN Dates might not be in history_tree display yet,
                # but are fetched by get_all_invoice_headers if DB was updated.
                # For display, ensure columns match or adapt here.
                display_values = (
                    header['invoice_no'], header['date'], header['customer_name'],
                    f"{header['total_amount']:.2f}",
                    f"{header.get('gst_percentage', 0.0):.2f}", # Use .get for new potentially missing cols
                    header.get('payment_method', ''),
                    # header.get('po_number', ''), # Example if adding to tree
                    # header.get('grn_date_from', ''),
                    # header.get('grn_date_to', '')
                )
                self.history_tree.insert("", tk.END, values=display_values[:len(self.history_tree["columns"])])

        else:
            if headers_to_load is not None:
                 print("No matching billing history found for the criteria.")
            else:
                 print("No billing history found or an error occurred during initial load.")

    def _search_invoices(self):
        search_term = self.search_invoice_no_var.get().strip()
        if not search_term:
            self._load_billing_history()
            return

        results = billing_db.get_all_invoice_headers(invoice_no_filter=search_term)
        if not results:
            messagebox.showinfo("Search Result", f"No invoices found matching '{search_term}'.")
        self._load_billing_history(headers_to_load=results)


# --- Existing BillingPreviewWindow Class (remains unchanged) ---
class BillingPreviewWindow(tk.Toplevel):
    def __init__(self, master, form_data, item_data):
        super().__init__(master)
        self.title("Challan Preview")
        self.geometry("1000x1000")
        self.configure(bg="#eeeeee")

        self.form_data = form_data
        self.item_data = item_data

        self.create_widgets()

    def create_widgets(self):
        preview_canvas = tk.Canvas(self, bg="#eeeeee")
        preview_canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=preview_canvas.yview)
        scrollbar.pack(side="right", fill="y")

        preview_canvas.configure(yscrollcommand=scrollbar.set)

        content_frame = tk.Frame(preview_canvas, bg="white")
        preview_canvas.create_window((0,0), window=content_frame, anchor="nw")

        content_frame.bind("<Configure>", lambda e: preview_canvas.configure(scrollregion=preview_canvas.bbox("all")))

        outer_big = tk.Frame(content_frame, bd=2, relief="solid", padx=10, pady=10, bg="white")
        outer_big.pack(padx=10, pady=10)

        outer = tk.Frame(outer_big, bd=2, relief="solid", padx=10, pady=10, bg="white")
        outer.pack(padx=5, pady=5)

        header_frame = tk.Frame(outer, bg="white")
        header_frame.pack(fill='x')
        tk.Label(header_frame, text="GSTIN : 33ALQPK2156A1ZG", font=("Arial", 10, "bold"), bg="white").pack(side="left")
        tk.Label(header_frame, text="Cell : 9443410161", font=("Arial", 10, "bold"), bg="white").pack(side="right")
        tk.Label(outer, text="Customer ID : 17124", font=("Arial", 9), bg="white").pack(anchor='w')

        outer_challan_frame = tk.Frame(outer, bg="white")
        outer_challan_frame.pack(pady=(10, 0))
        outer_box = tk.Frame(outer_challan_frame, bd=2, relief="solid", bg="white", padx=5, pady=5)
        outer_box.pack()
        inner_box = tk.Frame(outer_box, bd=1, relief="solid", bg="white", padx=10, pady=2)
        inner_box.pack()
        tk.Label(inner_box, text="DELIVERY CHALLAN", font=("Arial", 10, "bold"), bg="white").pack()

        tk.Label(outer, text="SRIRAM       COATERS", font=("Arial", 18, "bold"), bg="white").pack(pady=(10, 0))
        tk.Label(outer, text="Ambal Nagar Boothakudi Village, Viralimalai", font=("Arial", 10), bg="white").pack(pady=(0, 10))

        block_frame = tk.Frame(outer, bd=1, relief="solid", bg="white")
        block_frame.pack(fill="x", pady=5)
        left = tk.Frame(block_frame, bg="white")
        left.pack(side="left", fill="both", expand=True, padx=(10, 5), pady=5)
        tk.Label(left, text="TO,", font=("Arial", 8), bg="white").pack(anchor="w")
        tk.Label(left, text="M/s, ZF RANE AUTOMOTIVE INDIA PVT . LTD.,", font=("Arial", 10, "bold"), bg="white").pack(anchor="w")
        tk.Label(left, text="Boothakudi Village, Viralimalai - 621316", font=("Arial", 9), bg="white").pack(anchor="w")
        tk.Label(left, text="Pudukkottai District - 62316", font=("Arial", 9), bg="white").pack(anchor="w")
        tk.Label(left, text="GSTIN : 33AAACR3147C1ZY", font=("Arial", 8, "bold"), bg="white").pack(anchor="w")

        separator = tk.Frame(block_frame, bg="black", width=1)
        separator.pack(side="left", fill="y")

        right = tk.Frame(block_frame, bg="white")
        right.pack(side="left", fill="both", expand=True, padx=(5, 10), pady=5)
        tk.Label(right, text=f"DC.No.  : {self.form_data.get('DC No', '')}", font=("Arial", 12), bg="white").pack(anchor="w", padx=(15, 0))
        tk.Label(right, text=f"Date      : {self.form_data.get('Date (YYYY-MM-DD)', '')}", font=("Arial", 12), bg="white").pack(anchor="w", padx=(15, 0))

        doc_frame = tk.Frame(outer, bd=1, relief="solid", bg="white")
        doc_frame.pack(fill="x", pady=(5, 0))
        for label_text, value_key in [
            ("Po.No. :", "PO No"), ("DC.No. & Date", "DC No & Date"), ("CHALLAN No.", "Challan No")
        ]:
            f = tk.Frame(doc_frame, width=200, height=80, bd=1, relief="solid", bg="white")
            f.pack(side="left", expand=True, fill="both", ipady=10)
            tk.Label(f, text=label_text, font=("Arial", 10), bg="white").pack(anchor="n", pady=2)
            tk.Label(f, text=self.form_data.get(value_key, ""), font=("Arial", 10), bg="white").pack(anchor="center")

        table_frame = tk.Frame(outer, bd=1, relief="solid", bg="white")
        table_frame.pack(fill="x", pady=(10, 0))
        headers = ["Sl. No.", "SAC Code", "Customer Part No.", "Description", "Qty", "Rate", "Amount"]

        for col_index, header in enumerate(headers):
            cell = tk.Frame(table_frame, bd=1, relief="solid", bg="white", height=30)
            cell.grid(row=0, column=col_index, sticky="nsew")
            tk.Label(cell, text=header, font=("Arial", 9, "bold"), bg="white").pack(expand=True, fill="both")

        items_to_display = []
        if isinstance(self.item_data, list):
            for i, item in enumerate(self.item_data):
                 items_to_display.append([
                    str(i + 1), item.get("sac_code", ""), item.get("customer_part_no", ""),
                    item.get("description", ""), str(item.get("qty", "")),
                    str(item.get("rate", "")), str(item.get("amount", ""))
                 ])
        elif isinstance(self.item_data, dict):
            items_to_display.append([
                "1", self.item_data.get("sac_code", ""), self.item_data.get("customer_part_no", ""),
                self.item_data.get("description", ""), str(self.item_data.get("qty", "")),
                str(self.item_data.get("rate", "")), str(self.item_data.get("amount", ""))
            ])

        total_rows_to_display = max(8, len(items_to_display))

        for row_idx in range(total_rows_to_display):
            row_data = items_to_display[row_idx] if row_idx < len(items_to_display) else [""] * len(headers)
            for col_idx, val in enumerate(row_data):
                cell = tk.Frame(table_frame, bd=1, relief="solid", bg="white", height=40 if row_idx < len(items_to_display) else 30)
                cell.grid(row=row_idx + 1, column=col_idx, sticky="nsew")
                tk.Label(cell, text=val, bg="white", font=("Arial", 9)).pack(expand=True, fill="both", padx=2, pady=2)

        for col_index in range(len(headers)):
            table_frame.grid_columnconfigure(col_index, weight=1)

        declaration_frame = tk.Frame(outer, bd=1, relief="solid", bg="white")
        declaration_frame.pack(fill="x", pady=10)
        tk.Label(declaration_frame, text="Recieved the above Material in good condition",
                 font=("Arial", 10, "bold"), bg="white").pack(pady=(5, 2))

        bottom = tk.Frame(declaration_frame, bg="white")
        bottom.pack(fill="x", padx=10)
        left_txt = "We Hearby Declare that the above mentioned goods are returned\nafter processing and no material transactions are invloved."
        tk.Label(bottom, text=left_txt, font=("Arial", 8), bg="white", justify="left").pack(side="left", anchor="w")
        tk.Label(bottom, text="For SRIRAM COATERS", font=("Arial", 9, "bold"), bg="white").pack(side="right", anchor="e", pady=(5, 0))

        sign_frame = tk.Frame(declaration_frame, bg="white")
        sign_frame.pack(fill="x", pady=10)
        for label_text in ["Signature of Receiver", "Labour Charges Only", "Signature"]:
            l = tk.Label(sign_frame, text=label_text, font=("Arial", 9), bg="white")
            l.pack(side="left", expand=True)


if __name__ == '__main__':
    root = tk.Tk()
    root.title("Billing Entry Test")
    root.geometry("800x750")
    style = ttk.Style(root)
    style.configure("Accent.TButton", foreground="white", background="dodgerblue")
    billing_page = BillingEntryPage(root)
    billing_page.pack(fill="both", expand=True)

    root.mainloop()

